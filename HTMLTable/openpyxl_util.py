#!/usr/bin/env python
# -*- encoding=utf8 -*-

'''
Author: fasion
Created time: 2022-01-21 17:13:09
Last Modified by: fasion
Last Modified time: 2022-01-21 17:13:31
'''

from openpyxl.utils.cell import (
	column_index_from_string,
	coordinate_to_tuple,
	get_column_letter,
)

default_datetime_format_mapping = {
	'h:mm': '%H:%M',
	'mm-dd-yy': '%m-%d-%y'
}

def coordinates_of_cell_region(start, rows, columns):
	startColumn, startRow = coordinate_to_tuple(start)
	return tuple(
		'{:s}{:d}'.format(get_column_letter(column), row)
		for row in range(startRow, startRow+rows)
			for column in range(startColumn, startColumn+columns)
	)

def coordinates_of_merged_cell(merged):
	min_row, min_col = merged.min_row, merged.min_col
	max_row, max_col = merged.max_row, merged.max_col
	return tuple(
		'{:s}{:d}'.format(get_column_letter(col), row)
		for row in range(min_row, max_row+1)
			for col in range(min_col, max_col+1)
	)

def hidden_coordinates_of_merged_cell(merged):
	return coordinates_of_merged_cell(merged)[1:]

def hidden_coordinate_set_of_merged_cells(mergeds):
	s = set()
	for merged in mergeds:
		for coordinate in hidden_coordinates_of_merged_cell(merged):
			s.add(coordinate)
	return s

def all_cells_empty(cells):
	for cell in cells:
		if cell.value is not None:
			return False
	return True

def detect_sheet_end_column(sheet):
	index = sheet.max_column
	name = get_column_letter(index)
	while index > 1 and all_cells_empty(sheet[name]):
		index -= 1
		name = get_column_letter(index)
	return name

def merge_cell_mapping_by_top_left(ranges):
	return {
		'{:s}{:d}'.format(get_column_letter(item.min_col), item.min_row): item
		for item in ranges
	}

def fill_html_table_from_excel_sheet(table, sheet, start_row=1, end_row=None, start_column='A', end_column=None, ignoreEmptyRow=True, datetime_format_mapping=None):
	if not end_column:
		end_column = detect_sheet_end_column(sheet)

	start_column_index = column_index_from_string(start_column)
	end_column_index = column_index_from_string(end_column)

	hidden_cells = hidden_coordinate_set_of_merged_cells(sheet.merged_cells.ranges)
	merged_mapping = merge_cell_mapping_by_top_left(sheet.merged_cells.ranges)

	for ri, row in enumerate(sheet):
		ri += 1
		if ri < start_row:
			continue

		if end_row is not None and ri > end_row:
			continue

		if ignoreEmptyRow and all_cells_empty(row):
			continue

		html_row = table.append_row()

		for cell in row[start_column_index-1:end_column_index]:
			value = str(cell.value or '').replace('\n', '<br />')
			if cell.is_date and cell.value:
				fmt = (datetime_format_mapping or {}).get(cell.number_format)
				if not fmt:
					fmt = default_datetime_format_mapping.get(cell.number_format)
				if fmt:
					value = cell.value.strftime(fmt)

			html_cell = html_row.append_cell(value)
			html_cell.set_escape(False)

			merged = merged_mapping.get(cell.coordinate)
			if merged:
				html_cell.set_colspan(merged.max_col+1-merged.min_col)
				html_cell.set_rowspan(merged.max_row+1-merged.min_row)

			html_cell_style = {
				'text-align': cell.alignment.horizontal,
				'vertical-align': cell.alignment.vertical,
			}

			fill_color = cell.fill.start_color.index
			if fill_color and fill_color.startswith('FF'):
				html_cell_style['background-color'] = '#{:s}'.format(fill_color[2:])

			html_cell.set_style(html_cell_style)
